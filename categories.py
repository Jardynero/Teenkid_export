# Программа добавляет категории в колонку categories

from openpyxl import load_workbook
wb1 = load_workbook("Sizes/Размеры.xlsx")
ws1 = wb1.active


# Функция считывания колонки и добавление ее в список
def read_column(ws1, letter, empty_list):
    for data in ws1[letter]:
        empty_list.append(data.value)


vzroslie = []
igrushki = []
malishi = []
deti = []
podrostki = []
shapki = []
obuv_yaselnaya = []
obuv_malodetskaya = []
obuv_doshkolnaya = []
obuv_shkolnyay = []
obuv_zhenskaya = []
obuv_men = []


read_column(ws1=ws1, letter='A', empty_list=vzroslie)
read_column(ws1=ws1, letter='B', empty_list=igrushki)
read_column(ws1=ws1, letter='E', empty_list=malishi)
read_column(ws1=ws1, letter='F', empty_list=deti)
read_column(ws1=ws1, letter='G', empty_list=podrostki)
read_column(ws1=ws1, letter='H', empty_list=shapki)
read_column(ws1=ws1, letter='I', empty_list=obuv_yaselnaya)
read_column(ws1=ws1, letter='J', empty_list=obuv_malodetskaya)
read_column(ws1=ws1, letter='K', empty_list=obuv_doshkolnaya)
read_column(ws1=ws1, letter='L', empty_list=obuv_shkolnyay)
read_column(ws1=ws1, letter='M', empty_list=obuv_zhenskaya)
read_column(ws1=ws1, letter='N', empty_list=obuv_men)
wb1.close()


def first_categories(ws):
    row_counter = 1
    for row in ws['J']:
        row = row.value
        if row is None:
            row_counter += 1
            continue
        elif row in vzroslie:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Мужчинам'
            else:
                ws[f'E{row_counter}'].value = 'Женщинам'
        elif row in igrushki:
            ws[f'E{row_counter}'].value = 'Игрушки'
        elif row in malishi:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Малышам>Мальчикам'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Малышам>Девочкам'
        elif row in deti:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Мальчикам'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Девочкам'
        elif row in podrostki:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Мальчикам'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Девочкам'
        elif row in shapki:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Мальчикам'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Девочкам'
        elif row in obuv_yaselnaya:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Малышам>Мальчикам>Ясельная обувь'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Малышам>Девочкам>Ясельная обувь'
        elif row in obuv_malodetskaya:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Мальчикам>Малодетская обувь'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Девочкам>Малодетская обувь'
        elif row in obuv_doshkolnaya:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Мальчикам>Дошкольная обувь'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Девочкам>Дошкольная обувь'
        elif row in obuv_shkolnyay:
            if ws[f'N{row_counter}'].value == 'Мужской':
                ws[f'E{row_counter}'].value = 'Детям>Мальчикам>Школьная обувь'
            else:
                ws[f'E{row_counter}'].value = 'Детям>Девочкам>Школьная обувь'
        elif row in obuv_zhenskaya:
                ws[f'E{row_counter}'].value = 'Женщинам>Женская обувь'
        elif row in obuv_men:
            ws[f'E{row_counter}'].value = 'Мужчинам>Мужская обувь'
        row_counter += 1


# Функция добавляет подкатегории в колонку Categories
def sub_categories(ws):
    from openpyxl import load_workbook
    wb1 = load_workbook("Subcategories/subcategories.xlsx")
    ws1 = wb1.active

    odezhda = []
    belie = []
    aksessuari = []

    for data in ws1['A']:
        odezhda.append(data.value)

    for data in ws1['B']:
        belie.append(data.value)

    for data in ws1['C']:
        aksessuari.append(data.value)

    index_counter = 1
    for data in ws['D']:
        data = data.value
        current_value = ws[f'E{index_counter}'].value
        if data is None:
            index_counter += 1
            continue
        elif current_value is None:
            index_counter += 1
            continue
        elif data in odezhda:
            # ws[f'E{index_counter}'].value = current_value + '>' + 'Одежда'
            index_counter += 1
            continue
        elif data in belie:
            ws[f'E{index_counter}'].value = current_value + '>' + 'Белье и пижамы'
        elif data in aksessuari:
            ws[f'E{index_counter}'].value = current_value + '>' + 'Аксессуары'
        else:
            index_counter += 1
            continue
        index_counter += 1


# Функция править короткие наименования товаров
def catagories_names_repair(ws):
    index_counter = 0
    empty_list = []
    for data in ws['D']:
        empty_list.append(data.value)
    database = empty_list

    komplekt = ['Комплект', 'Комлект']
    futbolka = ['Футболка', 'Футболки']
    platie = ['Платье']
    vodolazka = ['Водолазка']
    dgemper = ['Джемпер мужской', 'Джемпер женский', 'Джемпер', 'Джемперы']
    blusa = ['Блуза', 'Блузка', 'Блузки']
    ubka = ['Юбка']
    kostum = ['Костюм']
    bodi = ['Боди-водолазка', 'Боди-Платье']
    kofta = ['Кофта', 'Кофточка', 'Кофточки', 'Кофта женская', 'Кофта для мальчика']
    rubashka = ['Рубашка', 'Рубашечка']
    kombinezon = ['Комбинезон', 'Комбинезон-трансформер', 'Полукомбинезон', 'Полукомбинзоны']
    kardigan = ['Кардиган', 'Кардиган (кофта) для девочки', 'Кардиган мужской', 'Кардиган женский']
    kurtka = ['Куртка', 'Куртка женская']
    zhilet = ['Жилет']
    tolstovka = ['Толстовка', 'Толстовка женская', 'Толстовка мужская']
    sarafan = ['Сарафан']
    trusi = ['Трусы-боксеры', 'Трусы-шорты', 'Трусы-Боксеры', 'Трусы-Шорты', 'Кальсоны', 'Трусы']
    maika = ['Майка', 'Майки бельевые', 'Майки']
    pizhama = ['Пижама']
    kupalnik = ['Купальник']
    noski = ['носки', 'Носки (3 шт.)', 'Носки детские']
    sorochki = ['Сорочка']
    kolgotki = ['колготки']
    halati = ['Халат']
    shapki = ['Шапка', 'Шапочка', 'Шапка-шлем', 'Шапка-Шлем', 'Головные уборы', 'Шапки']
    varezhki = ['Варежки детские']
    snud = ['Снуд']
    for data in database:
        if data is None:
            index_counter += 1
            continue
        elif data == 'Краткое наименование':
            index_counter += 1
            continue
        elif data in komplekt:
            database[index_counter] = 'Комплекты'
        elif data in futbolka:
            database[index_counter] = 'Футболки'
        elif data in platie:
            database[index_counter] = 'Платья'
        elif data in vodolazka:
            database[index_counter] = 'Водолазки'
        elif data in dgemper:
            database[index_counter] = 'Джемперы'
        elif data in blusa:
            database[index_counter] = 'Блузы'
        elif data in ubka:
            database[index_counter] = 'Юбки'
        elif data in kostum:
            database[index_counter] = 'Костюмы'
        elif data in bodi:
            database[index_counter] = 'Боди'
        elif data in kofta:
            database[index_counter] = 'Кофты'
        elif data in rubashka:
            database[index_counter] = 'Рубашки'
        elif data in kombinezon:
            database[index_counter] = 'Комбинезоны'
        elif data in kardigan:
            database[index_counter] = 'Кардиганы'
        elif data in kurtka:
            database[index_counter] = 'Куртки'
        elif data in zhilet:
            wdatabase[index_counter] = 'Жилеты'
        elif data in tolstovka:
            database[index_counter] = 'Толстовки'
        elif data in sarafan:
            database[index_counter] = 'Сарафаны'
        elif data in trusi:
            database[index_counter] = 'Трусы'
        elif data in maika:
            database[index_counter] = 'Майки'
        elif data in pizhama:
            database[index_counter] = 'Пижамы'
        elif data in kupalnik:
            database[index_counter] = 'Купальники'
        elif data in noski:
            database[index_counter] = 'Носки'
        elif data in sorochki:
            database[index_counter] = 'Сорочки'
        elif data in kolgotki:
            database[index_counter] = 'Колготки'
        elif data in halati:
            database[index_counter] = 'Халаты'
        elif data in shapki:
            database[index_counter] = 'Шапки'
        elif data in varezhki:
            database[index_counter] = 'Варежки'
        elif data in snud:
            database[index_counter] = 'Снуды'
        index_counter += 1
    index_counter = 1
    for data in database:
        ws[f'D{index_counter}'].value = data
        index_counter += 1

# Функция меняет short title(сокращение категорий)
def edit_shorttitle(ws):
    kofty = ['Болеро', 'Джемперы', 'Кардиганы', 'Лонгсливы',
            'Свитшоты', 'Толстовки', 'Кофты', 'Бомберы', 'Куртки']
    leggincy_bridgi = ['Лосины', 'Леггинсы', 'Бриджи', 'Легинсы']
    maika = ['Майка', 'Майки бельевые', 'Майки']
    index_counter = 1
    for data in ws['D']:
        data = data.value
        if data in kofty:
            ws[f'D{index_counter}'].value = "Кофты"
            index_counter += 1
        elif data in leggincy_bridgi:
            ws[f'D{index_counter}'].value = 'Леггинсы и бриджи'
            index_counter += 1
        elif data == "Футболки":
            ws[f'D{index_counter}'].value = "Футболки и майки"
            index_counter += 1
        elif data == "Майки":
            ws[f'D{index_counter}'].value = "Футболки и майки"
            index_counter += 1
        else:
            index_counter += 1

# Функция добавяет последнюю часть категорий в колонку Categories
def end_categories(ws):
    index_counter = 1
    for data in ws['D']:
        current_value = ws[f'E{index_counter}'].value
        if data.value is None:
            index_counter += 1
            continue
        elif current_value is None:
            index_counter += 1
            continue
        else:
            ws[f'E{index_counter}'].value = current_value + '>' + data.value
        index_counter += 1


# Функция вносит корректировки в названия 
def correction_in_titles(ws, categorie, current_title, new_title):
    index_counter = 1
    for data in ws['E']:
        data = data.value
        if data != categorie:
            index_counter += 1
            continue
        elif data == categorie:
            title = ws[f'F{index_counter}'].value
            if title.find(current_title) > 0:
                title = title.replace(current_title, new_title)
                ws[f'F{index_counter}'].value = title
            index_counter += 1
