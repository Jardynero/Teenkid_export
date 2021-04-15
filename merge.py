import os
from openpyxl import load_workbook

# Открыть ексель с базой
wb_database = load_workbook('database/database.xlsx')
ws_database = wb_database.active

# Открыть ексель с новыми товарами
wb_new = load_workbook('new_product.xlsx')
ws_new = wb_new.active


# Функция добавления списка артикулей из файла в список
def add_sku_to_list(ws, letter, list_to_append):
	for sku in ws[letter]:
		list_to_append.append(sku.value)
	return set(list_to_append)


# пустые списки
new = []
database = []

# Получение списков с артикулами из файла
new_sku = add_sku_to_list(ws=ws_new, letter='A', list_to_append=new)
database_sku = add_sku_to_list(ws=ws_database, letter='A', list_to_append=database)

# Возвращает номер последней строки
def len_sku_col():
	last_row = len(ws_database['A'])
	return last_row + 1

def add_new_products():
	list_with_rows = []
	index_counter = 1
	new_product_sku = new_sku - database_sku
	for sku in new_product_sku:
		index_counter = 1
		for data in ws_new['A']:
			if sku == data.value:
				cell_range = ws_new[f'A{index_counter}:X{index_counter}']
				last_row = len_sku_col()
				for product in cell_range:
					counter = 1
					for values in product:
						ws_database.cell(row=last_row, column=counter, value=values.value)
						counter += 1
				list_with_rows.append(index_counter)
			index_counter += 1
	print(f'Было добавленно {len(new_product_sku)} новых товаров')


def update():
	index_counter_1 = 1
	for data_new in ws_new['A']:
		index_counter_2 = 1
		for data_database in ws_database['A']:
			if data_new.value == data_database.value:
				if ws_new[f'H{index_counter_1}'].value == ws_database[f'H{index_counter_2}'].value:
					if ws_new[f'J{index_counter_1}'].value == ws_database[f'J{index_counter_2}'].value:
						stock_value = ws_new[f'K{index_counter_1}'].value
						print(stock_value)
						ws_database[f'K{index_counter_2}'].value = stock_value
			index_counter_2 += 1
		index_counter_1 += 1


# Убрать всем товаров наличие(ставится 0)
for stock in ws_database['K']:
	if stock.value == 'stock':
		continue
	stock.value = '0'

# ws_database.cell(row=len_sku_col(), column=1, value=10)
# cell_range = ws_new['A1':'X1']
add_new_products()
update()
wb_database.save('database/database.xlsx')
