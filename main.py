import os
import pyexcel
import sys
import excel_file_search
from delete_cols import delete_cols
from xlsx_to_csv import convert_excel_to_csv
from articles import sortexcel, make_articles
from timeit import default_timer as timer
from openpyxl import load_workbook
import time
import datetime
from insert_cols import inserting_cols, name_columns
from photos import *
from price import *
from titles import title_correction
from titles import second_title_correction_e
from titles import second_title_correction_f
from filter_color import filter_color
from categories import *
from featured import featured_products

# текущее время
localtime = time.localtime()

# Запустим таймер для проверки затраты скрипта
start_time = timer()

# Проверка на количество файлов в корневом каталоге
excel_file_search.number_of_xlsx_files()
# Определение название файла. По необзодимости конвертация в .xlsx или выход
excel_file_search.detect_file()
# Сортировка файла от "а до я" с помощью pandas
sortexcel()

# Инициализирую ексель файл
wb = load_workbook(excel_file_search.excelfilename())
ws = wb.active

# Функция производит изменения в Артикулах
make_articles(wb=wb, ws=ws)
# Функуция убирает артикулы в начале строк в столбцах E и F
title_correction(ws=ws)
# Функция убирает оставшиеся скобочки вначале строк столбца "наименование"
second_title_correction_e(ws=ws)
# Функция убирает оставшиеся скобочки вначале строк столбца "полное наименование"
second_title_correction_f(ws=ws)
# Функция удаляет ненужные колонки
delete_cols(wb=wb, ws=ws)
# Функция добавляет новые колонки в таблицу
inserting_cols(ws=ws)
# Функция дает названия новым колонкам
name_columns(ws=ws)
# Функция добавляет названия фотографий товаров в колонку "С"
add_photo_titles(ws=ws)
# Функция делает надбавку у цены и вписывает ее в колонку "M"
add_extra_charge(ws=ws)
# Функция добавляет цвета в колонку Filter color
filter_color(ws=ws)
# Функция правит наименования
catagories_names_repair(ws=ws)
# Функция меняет short title(сокращение категорий)
edit_shorttitle(ws=ws)
# Функция проставляет первую часть категорий
first_categories(ws=ws)
# Функция добавляет подкатегории в колонку Categories
sub_categories(ws=ws)
# Функция добавяет последнюю часть категорий в колонку Categories
end_categories(ws=ws)
# Функция добавляет рекомендованные продукты
featured_products(ws=ws)
add_discount(ws=ws, articles=['CSJG 62673', 
								'CSJG 62666', 
								'CWJ 7963', 
								'CSJB 62650', 
								'FSSW 60037', 
								'CSJB 62651', 
								'CSJB 62730', 
								'CSJB 62733',
								'CWK 62504',
								'CWK 62301'])
# Функция добавляет название фотографий с сайта черубино
add_cherubino_photo(ws, os)


def new_articles(ws):
	ws['A1'].value = 'main sku'
	ws['B1'].value = 'sku'
	ws['C1'].value = 'photos'
	ws['D1'].value = 'short title'
	ws['E1'].value = 'categories'
	ws['F1'].value = 'title'
	ws['G1'].value = 'full description'
	ws['H1'].value = 'color'
	ws['I1'].value = 'filter color'
	ws['J1'].value = 'size'
	ws['K1'].value = 'stock'
	ws['L1'].value = 'Price NOT'
	ws['M1'].value = 'Price with extra charge'
	ws['N1'].value = 'Price with discount'
	ws['O1'].value = 'sex'
	ws['P1'].value = 'material'
	ws['R1'].value = 'country'
	ws['S1'].value = 'brand'
	ws['W1'].value = 'featured'
	ws['X1'].value = 'collections'

# Меняю заголовки у столбцов
new_articles(ws=ws)

# Получение названия файла ексель
current_excel_file_name = excel_file_search.excelfilename()
# Константа - название для сохранения выгрузки.
XLSX_FILE_NAME = f'Выгрузка Teenkid от {datetime.date.today()}|{localtime.tm_hour}:{localtime.tm_min}.xlsx'

# Сохранение файла с расширением .xlsx
print('Соранение файла выгрузки с расширением .xlsx')
wb.save(filename=XLSX_FILE_NAME)
wb.close()
print('Файл выгрузки успешно сохранен!')

# Удаление дефолтной выгрузки(предоставленной поставщиков Teenkid
os.remove(current_excel_file_name)

# Конвертирую в csv
convert_excel_to_csv()
print('На данный скрипт потраченно: {} времени'.format(timer() - start_time))
