# Программа добавляет колонку Filter Color


# Функция добавления данных из экселя в список
def obtain_color(ws1, letter, color_list):
    for i in ws1[letter]:
        color_list.append(i.value)


# Получаю все цвета из базы
def filter_color(ws):
    print('Начинаю фильтровать цвета')
    from openpyxl import load_workbook
    wb1 = load_workbook('Color/Цвета1.xlsx')
    ws1 = wb1.active

    # Пустые списки для цветов
    RED = []
    ORANGE = []
    YELLOW = []
    GREEN = []
    BLUE_CLAIR = []
    BLUE = []
    VIOLET = []
    WHITE = []
    BLACK = []
    BROWN = []
    GRAY = []
    PINK = []

    # Считываю цвета из базы
    print('Считываю базу цветов')
    obtain_color(ws1=ws1, letter='A', color_list=RED)
    obtain_color(ws1=ws1, letter='B', color_list=ORANGE)
    obtain_color(ws1=ws1, letter='C', color_list=YELLOW)
    obtain_color(ws1=ws1, letter='D', color_list=GREEN)
    obtain_color(ws1=ws1, letter='E', color_list=BLUE_CLAIR)
    obtain_color(ws1=ws1, letter='F', color_list=BLUE)
    obtain_color(ws1=ws1, letter='G', color_list=VIOLET)
    obtain_color(ws1=ws1, letter='H', color_list=WHITE)
    obtain_color(ws1=ws1, letter='I', color_list=BLACK)
    obtain_color(ws1=ws1, letter='J', color_list=BROWN)
    obtain_color(ws1=ws1, letter='K', color_list=GRAY)
    obtain_color(ws1=ws1, letter='L', color_list=PINK)

    # Проставляю цвета в колонку Filter color
    print('Добавляю цвета в Filter color')
    row_counter = 1
    for color in ws['H']:
        color = color.value
        if color is None:
            row_counter += 1
            continue
        elif color.lower() == 'Цвет':
            row_counter += 1
            continue
        elif color.lower() in RED:
            ws[f'I{row_counter}'].value = 'Красный'
        elif color.lower() in ORANGE:
            ws[f'I{row_counter}'].value = 'Оранжевый'
        elif color.lower() in YELLOW:
            ws[f'I{row_counter}'].value = 'Желтый'
        elif color.lower() in GREEN:
            ws[f'I{row_counter}'].value = 'Зеленый'
        elif color.lower() in BLUE_CLAIR:
            ws[f'I{row_counter}'].value = 'Голубой'
        elif color.lower() in BLUE:
            ws[f'I{row_counter}'].value = 'Синий'
        elif color.lower() in VIOLET:
            ws[f'I{row_counter}'].value = 'Фиолетовы'
        elif color.lower() in WHITE:
            ws[f'I{row_counter}'].value = 'Белый'
        elif color.lower() in BLACK:
            ws[f'I{row_counter}'].value = 'Черный'
        elif color.lower() in BROWN:
            ws[f'I{row_counter}'].value = 'Коричневый'
        elif color.lower() in GRAY:
            ws[f'I{row_counter}'].value = 'Серый'
        elif color.lower() in PINK:
            ws[f'I{row_counter}'].value = 'Розовый'
        row_counter += 1
    print("Фильтрация цветов завершенна")