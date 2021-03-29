from openpyxl import load_workbook
from excel_file_search import excelfilename

wb = load_workbook('Sizes/Размеры.xlsx')
ws = wb.active

SIZES_FROM_DATABASE = []
SIZES_FROM_TEENKID_FILE = []
SIZES_NOT_IN_DATABASE = []
for i in ws['A1:D259']:
    for cell in i:
        SIZES_FROM_DATABASE.append(cell.value)

wb = load_workbook(excelfilename())
ws = wb.active

for i in ws['J']:
    value = i.value
    if value is None:
        continue
    else:
        SIZES_FROM_TEENKID_FILE.append(value.lower())

SIZES_NOT_IN_DATABASE = list(set(SIZES_FROM_TEENKID_FILE) - set(SIZES_FROM_DATABASE))
f = open("Sizes/размеры не в базе.txt", 'w')
for i in SIZES_NOT_IN_DATABASE:
    f.write(i + '\n')
f.close()
